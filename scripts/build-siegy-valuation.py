#!/usr/bin/env python3
"""Build the filing-backed Siemens AG ADR normalized earnings-power receipt."""
from __future__ import annotations

import argparse
import json
import zipfile
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "trading" / "research"
STEM = "siegy-normalized-earnings-2026-08-07"


def scenario(normalized_eps_eur: float, multiple: float, fx: float = 1.16) -> dict[str, float]:
    ordinary_value_eur = normalized_eps_eur * multiple
    adr_value_usd = ordinary_value_eur * fx / 2.0
    return {
        "normalized_eps_eur_per_ordinary_share": normalized_eps_eur,
        "earnings_multiple": multiple,
        "ordinary_value_eur": ordinary_value_eur,
        "fair_value_per_share": round(adr_value_usd, 2),
    }


def build_payload() -> dict[str, Any]:
    quote = 161.95
    fx = 1.16
    base_multiple = 22.0
    return {
        "symbol": "SIEGY",
        "company": "Siemens AG ADR",
        "valuation_date": "2026-08-07",
        "method": "ADR-adjusted cycle-normalized earnings power",
        "currency": "USD per SIEGY ADR",
        "adr_terms": {
            "adr_per_ordinary_share": 2,
            "source": "Siemens official ADR page",
        },
        "market_snapshot": {
            "quote_usd": quote,
            "market_cap_usd_b": 246.897,
            "trailing_pe": 28.64,
            "dividend_yield_pct": 1.40,
            "source": "Robinhood regular-session data through 2026-08-07",
        },
        "reported_inputs": {
            "q3_revenue_eur_b": 20.794,
            "q3_orders_eur_b": 27.902,
            "q3_industrial_business_margin_pct": 17.3,
            "q3_free_cash_flow_eur_b": 4.148,
            "q1_q3_revenue_eur_b": 59.688,
            "q1_q3_free_cash_flow_eur_b": 6.541,
            "order_backlog_eur_b": 132.0,
            "fy2026_eps_pre_ppa_guide_low_eur": 11.20,
            "fy2026_eps_pre_ppa_guide_high_eur": 11.50,
        },
        "model_assumptions": {
            "eurusd": fx,
            "fx_note": "Static scenario translation assumption, not an FX forecast",
            "base_multiple": base_multiple,
        },
        "scenarios": {
            "bear": scenario(9.0, 18.0, fx),
            "base": scenario(12.0, base_multiple, fx),
            "bull": scenario(15.0, 26.0, fx),
        },
        "market_implied": {
            "normalized_eps_eur_at_base_multiple": round(quote * 2.0 / fx / base_multiple, 2),
            "premium_to_fy2026_eps_pre_ppa_midpoint_pct": round(
                (quote * 2.0 / fx / base_multiple) / 11.35 - 1.0,
                4,
            ),
        },
        "sources": [
            {
                "label": "Siemens Q3 FY2026 earnings release",
                "url": "https://assets.new.siemens.com/siemens/assets/api/uuid:a659b574-caac-4ea1-a94f-a9d624fe0b7f/2026-q3-earnings-release-en.pdf",
                "use": "Q3 orders, revenue, margin, cash flow, backlog and FY2026 outlook",
            },
            {
                "label": "Siemens financial calendar",
                "url": "https://www.siemens.com/en-us/company/investor-relations/financial-calendar/",
                "use": "Official November 12, 2026 Q4/FY2026 results date",
            },
            {
                "label": "Siemens ADR terms",
                "url": "https://www.siemens.com/en-us/company/investor-relations/share-bonds-rating/american-depository-receipt/",
                "use": "Two SIEGY ADRs represent one Siemens ordinary share",
            },
        ],
        "limitations": [
            "This is normalized earnings power, not a full sum-of-the-parts valuation.",
            "EPS pre PPA is a company-defined non-GAAP measure; acquisition accounting and restructuring remain economic costs.",
            "The static EUR/USD translation can move ADR value even when the underlying euro thesis is unchanged.",
            "The planned Siemens Healthineers spin-off can change both earnings composition and the appropriate multiple.",
            "SIEGY trades OTC with materially lower liquidity than the German ordinary shares.",
        ],
    }


def clean(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 8)
    if isinstance(value, dict):
        return {key: clean(item) for key, item in value.items()}
    if isinstance(value, list):
        return [clean(item) for item in value]
    return value


def style_sheet(ws: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(wrap_text=True)
    for column in range(1, ws.max_column + 1):
        width = max(len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 72)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(payload: dict[str, Any], path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["Siemens AG ADR earnings-power receipt", "Value"])
    summary.append(["Valuation date", payload["valuation_date"]])
    summary.append(["Quote (USD/ADR)", payload["market_snapshot"]["quote_usd"]])
    summary.append(["Method", payload["method"]])
    summary.append(["Bear / Base / Bull", " / ".join(f"${payload['scenarios'][case]['fair_value_per_share']:.2f}" for case in ("bear", "base", "bull"))])
    summary.append(["Market-implied normalized EPS (€)", payload["market_implied"]["normalized_eps_eur_at_base_multiple"]])
    style_sheet(summary)

    inputs = wb.create_sheet("Inputs")
    inputs.append(["Input", "Value", "Unit / note"])
    for key, value in payload["reported_inputs"].items():
        inputs.append([key, value, "Siemens Q3 FY2026 release"])
    for key, value in payload["model_assumptions"].items():
        inputs.append([key, value, "Model assumption"])
    style_sheet(inputs)

    for name in ("bear", "base", "bull"):
        case = payload["scenarios"][name]
        ws = wb.create_sheet(name.title())
        ws.append(["Normalized EPS (€ / ordinary)", "Multiple", "Ordinary value (€)", "ADR ratio", "EUR/USD", "ADR value (USD)"])
        ws.append([case["normalized_eps_eur_per_ordinary_share"], case["earnings_multiple"], "=A2*B2", 2, payload["model_assumptions"]["eurusd"], "=C2/D2*E2"])
        ws.append(["Receipt output", "", case["ordinary_value_eur"], "", "", case["fair_value_per_share"]])
        style_sheet(ws)

    sources = wb.create_sheet("Sources")
    sources.append(["Source", "URL", "Use"])
    for source in payload["sources"]:
        sources.append([source["label"], source["url"], source["use"]])
        sources.cell(sources.max_row, 2).hyperlink = source["url"]
        sources.cell(sources.max_row, 2).style = "Hyperlink"
    style_sheet(sources)

    limitations = wb.create_sheet("Limitations")
    limitations.append(["Model limitation", "Implication"])
    for item in payload["limitations"]:
        limitations.append([item, "Review before relying on scenario output"])
    style_sheet(limitations)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_all() -> None:
    payload = clean(build_payload())
    OUT.mkdir(parents=True, exist_ok=True)
    json_path = OUT / f"{STEM}.json"
    xlsx_path = OUT / f"{STEM}.xlsx"
    json_path.write_text(json.dumps(payload, indent=2) + "\n")
    build_workbook(payload, xlsx_path)
    print(f"wrote {json_path.relative_to(ROOT)} and {xlsx_path.relative_to(ROOT)}")


def check_all() -> None:
    payload = clean(build_payload())
    json_path = OUT / f"{STEM}.json"
    xlsx_path = OUT / f"{STEM}.xlsx"
    if json.loads(json_path.read_text()) != payload:
        raise SystemExit(f"stale receipt: {json_path}")
    if xlsx_path.stat().st_size < 10_000 or not zipfile.is_zipfile(xlsx_path):
        raise SystemExit(f"invalid workbook: {xlsx_path}")
    with zipfile.ZipFile(xlsx_path) as workbook:
        workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8")
        for sheet in ("Summary", "Inputs", "Bear", "Base", "Bull", "Sources", "Limitations"):
            if f'name="{sheet}"' not in workbook_xml:
                raise SystemExit(f"workbook missing {sheet}")
        base_xml = workbook.read("xl/worksheets/sheet4.xml").decode("utf-8")
        if "<f>A2*B2</f>" not in base_xml or "<f>C2/D2*E2</f>" not in base_xml:
            raise SystemExit("workbook lost scenario formulas")
    print("SIEGY valuation receipt is current")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    check_all() if args.check else write_all()


if __name__ == "__main__":
    main()
