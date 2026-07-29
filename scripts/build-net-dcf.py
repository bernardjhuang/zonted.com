#!/usr/bin/env python3
"""Build Cloudflare's filing-backed 10-year DCF workbook and machine receipt."""
from __future__ import annotations

import argparse
import json
import math
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "trading" / "research"
JSON_OUT = OUT_DIR / "net-dcf-2026-07-29.json"
XLSX_OUT = OUT_DIR / "net-dcf-2026-07-29.xlsx"

VALUATION_DATE = "2026-07-29"
QUOTE = 265.995
QUOTE_TIMESTAMP = "2026-07-29T13:55:23.901427449Z"
REMAINING_YEAR_FRACTION = 0.425
FY2025_REVENUE = 2.1679  # $B
FY2026_REVENUE_GUIDE = 2.8090  # $B, midpoint of $2.805B-$2.813B
CASH_AND_SECURITIES = 4.163878  # $B at 2026-03-31
DILUTED_SHARES = 0.375  # billions; FY2026 management guidance, as-converted treatment
RESTRUCTURING_CASH = 0.1075  # $B midpoint; majority expected in Q2 2026
TTM_REVENUE = 2.328568  # $B: FY2025 + Q1'26 - Q1'25
TTM_FCF = 0.291769  # $B: management FCF, same bridge
TTM_SBC = 0.470160  # $B GAAP share-based compensation, same bridge

SCENARIOS = {
    "bear": {
        "revenue_growth": [None, 0.22, 0.18, 0.15, 0.13, 0.11, 0.09, 0.07, 0.055, 0.045],
        "fcf_margin": [0.11, 0.12, 0.14, 0.16, 0.18, 0.19, 0.20, 0.21, 0.22, 0.23],
        "discount_rate": 0.125,
        "terminal_growth": 0.025,
        "annual_dilution": 0.012,
    },
    "base": {
        "revenue_growth": [None, 0.25, 0.22, 0.19, 0.16, 0.13, 0.11, 0.09, 0.07, 0.05],
        "fcf_margin": [0.13, 0.15, 0.18, 0.21, 0.23, 0.25, 0.26, 0.27, 0.28, 0.28],
        "discount_rate": 0.11,
        "terminal_growth": 0.03,
        "annual_dilution": 0.01,
    },
    "bull": {
        "revenue_growth": [None, 0.28, 0.25, 0.22, 0.19, 0.16, 0.13, 0.11, 0.085, 0.06],
        "fcf_margin": [0.14, 0.17, 0.21, 0.25, 0.28, 0.30, 0.32, 0.33, 0.34, 0.34],
        "discount_rate": 0.095,
        "terminal_growth": 0.035,
        "annual_dilution": 0.008,
    },
}

SOURCES = [
    {
        "label": "Cloudflare Q1 2026 Form 10-Q",
        "url": "https://www.sec.gov/Archives/edgar/data/1477333/000147733326000038/cloud-20260331.htm",
        "use": "Revenue, FCF, RPO, customer metrics, cash/securities, converts, SBC, restructuring",
    },
    {
        "label": "Cloudflare Q1 2026 earnings release (8-K/A exhibit)",
        "url": "https://www.sec.gov/Archives/edgar/data/1477333/000147733326000035/q126exhibit991.htm",
        "use": "FY2026 revenue guidance and restructuring cash range",
    },
    {
        "label": "Cloudflare 2025 Form 10-K",
        "url": "https://www.sec.gov/Archives/edgar/data/1477333/000147733326000016/cloud-20251231.htm",
        "use": "FY2025 revenue, cash flow, capex, SBC and capital structure",
    },
    {
        "label": "Robinhood NET quote and fundamentals",
        "url": "https://robinhood.com/stocks/NET",
        "use": "Timestamped quote and market capitalization",
    },
]


def forecast(name: str, spec: dict | None = None) -> dict:
    spec = dict(spec or SCENARIOS[name])
    revenue = FY2026_REVENUE_GUIDE
    rows = []
    explicit_pv_per_share = 0.0
    for index, year in enumerate(range(2026, 2036)):
        growth = spec["revenue_growth"][index]
        if index:
            revenue *= 1 + growth
        fcf = revenue * spec["fcf_margin"][index]
        if year == 2026:
            fcf -= RESTRUCTURING_CASH
        shares = DILUTED_SHARES * (1 + spec["annual_dilution"]) ** index
        discount_period = REMAINING_YEAR_FRACTION + index
        discount_factor = (1 + spec["discount_rate"]) ** discount_period
        pv_per_share = (fcf / shares) / discount_factor
        explicit_pv_per_share += pv_per_share
        rows.append(
            {
                "year": year,
                "revenue_growth": growth,
                "revenue_b": revenue,
                "fcf_margin": spec["fcf_margin"][index],
                "restructuring_cash_b": RESTRUCTURING_CASH if year == 2026 else 0.0,
                "fcf_b": fcf,
                "diluted_shares_b": shares,
                "discount_period": discount_period,
                "pv_per_share": pv_per_share,
            }
        )
    last = rows[-1]
    terminal_value_per_share = (
        last["fcf_b"]
        * (1 + spec["terminal_growth"])
        / (spec["discount_rate"] - spec["terminal_growth"])
        / last["diluted_shares_b"]
    )
    terminal_pv_per_share = terminal_value_per_share / (
        (1 + spec["discount_rate"]) ** last["discount_period"]
    )
    cash_per_share = CASH_AND_SECURITIES / DILUTED_SHARES
    fair_value = explicit_pv_per_share + terminal_pv_per_share + cash_per_share
    return {
        "name": name,
        "assumptions": spec,
        "rows": rows,
        "explicit_pv_per_share": explicit_pv_per_share,
        "terminal_pv_per_share": terminal_pv_per_share,
        "cash_per_share": cash_per_share,
        "fair_value_per_share": fair_value,
        "terminal_value_share_of_operating_value": terminal_pv_per_share
        / (explicit_pv_per_share + terminal_pv_per_share),
        "upside_downside_to_quote": fair_value / QUOTE - 1,
    }


def base_value(discount_rate: float, terminal_growth: float) -> float:
    spec = dict(SCENARIOS["base"])
    spec["discount_rate"] = discount_rate
    spec["terminal_growth"] = terminal_growth
    return forecast("base sensitivity", spec)["fair_value_per_share"]


def reverse_value(constant_growth: float) -> tuple[float, float, float]:
    spec = dict(SCENARIOS["base"])
    spec["revenue_growth"] = [None] + [constant_growth] * 9
    result = forecast("reverse growth", spec)
    last = result["rows"][-1]
    return result["fair_value_per_share"], last["revenue_b"], last["fcf_b"]


def solve_reverse_growth() -> dict:
    low, high = 0.0, 1.0
    for _ in range(100):
        mid = (low + high) / 2
        value, _, _ = reverse_value(mid)
        if value < QUOTE:
            low = mid
        else:
            high = mid
    growth = (low + high) / 2
    value, revenue, fcf = reverse_value(growth)
    return {
        "constant_revenue_growth_2027_2035": growth,
        "implied_2035_revenue_b": revenue,
        "implied_2035_fcf_b": fcf,
        "matched_value_per_share": value,
        "held_fixed": "2026 guide; base FCF-margin path ending at 28%; 11% discount rate; 3% terminal growth; 1% annual dilution",
    }


def solve_reverse_margin() -> dict:
    growth_path = SCENARIOS["base"]["revenue_growth"]

    def value(terminal_margin: float) -> tuple[float, float]:
        spec = dict(SCENARIOS["base"])
        spec["revenue_growth"] = growth_path
        spec["fcf_margin"] = [
            0.13 + (terminal_margin - 0.13) * index / 9 for index in range(10)
        ]
        result = forecast("reverse margin", spec)
        return result["fair_value_per_share"], result["rows"][-1]["fcf_b"]

    low, high = 0.1, 3.0
    for _ in range(100):
        mid = (low + high) / 2
        matched, _ = value(mid)
        if matched < QUOTE:
            low = mid
        else:
            high = mid
    margin = (low + high) / 2
    matched, fcf = value(margin)
    return {
        "terminal_fcf_margin": margin,
        "implied_2035_fcf_b": fcf,
        "matched_value_per_share": matched,
        "held_fixed": "base revenue path; 11% discount rate; 3% terminal growth; 1% annual dilution",
    }


def heroic_case() -> dict:
    spec = {
        "revenue_growth": [None] + [0.30] * 9,
        "fcf_margin": [0.14 + (0.40 - 0.14) * index / 9 for index in range(10)],
        "discount_rate": 0.10,
        "terminal_growth": 0.035,
        "annual_dilution": 0.008,
    }
    result = forecast("heroic market case", spec)
    last = result["rows"][-1]
    return {
        "fair_value_per_share": result["fair_value_per_share"],
        "2035_revenue_b": last["revenue_b"],
        "2035_fcf_b": last["fcf_b"],
        "assumptions": spec,
    }


def build_payload() -> dict:
    scenarios = {name: forecast(name) for name in SCENARIOS}
    discount_rates = [0.095, 0.10, 0.105, 0.11, 0.115, 0.12, 0.125]
    terminal_growth_rates = [0.02, 0.025, 0.03, 0.035, 0.04]
    sensitivity = {
        f"{discount_rate:.3f}": {
            f"{terminal_growth:.3f}": base_value(discount_rate, terminal_growth)
            for terminal_growth in terminal_growth_rates
        }
        for discount_rate in discount_rates
    }
    return {
        "schema_version": 1,
        "symbol": "NET",
        "valuation_date": VALUATION_DATE,
        "quote": QUOTE,
        "quote_timestamp": QUOTE_TIMESTAMP,
        "method": "SEC-first 10-year diluted per-share FCF DCF with a mid-year stub",
        "capital_structure_treatment": "As-converted: use management's 375M FY2026 diluted-share guide, ignore capped-call offsets, and do not subtract convertible principal again; add cash and securities.",
        "reported_inputs_b": {
            "fy2025_revenue": FY2025_REVENUE,
            "fy2026_revenue_guide_midpoint": FY2026_REVENUE_GUIDE,
            "cash_and_securities": CASH_AND_SECURITIES,
            "restructuring_cash_midpoint": RESTRUCTURING_CASH,
            "ttm_revenue": TTM_REVENUE,
            "ttm_free_cash_flow": TTM_FCF,
            "ttm_stock_based_compensation": TTM_SBC,
        },
        "scenarios": scenarios,
        "sensitivity": sensitivity,
        "reverse_dcf_growth": solve_reverse_growth(),
        "reverse_dcf_margin": solve_reverse_margin(),
        "heroic_market_case": heroic_case(),
        "sources": SOURCES,
        "limitations": [
            "Reported FCF adds back stock compensation. The model represents that economic cost through explicit annual dilution instead of subtracting SBC twice.",
            "The 2026 restructuring cash charge is deducted once at its disclosed midpoint.",
            "The 375M share guide is an as-converted treatment; capped-call offsets are ignored conservatively.",
            "Q2 2026 earnings on August 6 can make the near-term revenue and margin path stale immediately.",
        ],
    }


def style_sheet(sheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    navy = "15212B"
    blue = "2F80ED"
    pale = "EAF2F8"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    for column in range(1, sheet.max_column + 1):
        width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 52)
    for row in range(2, sheet.max_row + 1):
        if row % 2 == 0:
            for cell in sheet[row]:
                cell.fill = PatternFill("solid", fgColor=pale)
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.sheet_format.defaultRowHeight = 18


def build_workbook(payload: dict) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(["Cloudflare (NET) DCF", "Value"])
    rows = [
        ("Valuation date", payload["valuation_date"]),
        ("Quote", payload["quote"]),
        ("Quote timestamp", payload["quote_timestamp"]),
        ("Method", payload["method"]),
        ("Capital structure", payload["capital_structure_treatment"]),
        ("TTM revenue ($B)", TTM_REVENUE),
        ("TTM reported FCF ($B)", TTM_FCF),
        ("TTM SBC ($B)", TTM_SBC),
        ("Cash + securities ($B)", CASH_AND_SECURITIES),
    ]
    for name in ("bear", "base", "bull"):
        result = payload["scenarios"][name]
        rows.extend(
            [
                (f"{name.title()} fair value", result["fair_value_per_share"]),
                (f"{name.title()} upside/downside", result["upside_downside_to_quote"]),
                (f"{name.title()} terminal-value share", result["terminal_value_share_of_operating_value"]),
            ]
        )
    rows.extend(
        [
            ("Reverse DCF: required annual revenue growth, 2027-35", payload["reverse_dcf_growth"]["constant_revenue_growth_2027_2035"]),
            ("Reverse DCF: implied 2035 revenue ($B)", payload["reverse_dcf_growth"]["implied_2035_revenue_b"]),
            ("Reverse DCF: required 2035 FCF margin on base growth", payload["reverse_dcf_margin"]["terminal_fcf_margin"]),
            ("Heroic case fair value", payload["heroic_market_case"]["fair_value_per_share"]),
            ("Heroic case 2035 revenue ($B)", payload["heroic_market_case"]["2035_revenue_b"]),
            ("Heroic case 2035 FCF ($B)", payload["heroic_market_case"]["2035_fcf_b"]),
        ]
    )
    for row in rows:
        summary.append(row)
    for row in summary.iter_rows(min_row=2):
        label = str(row[0].value)
        if "upside/downside" in label or "share" in label or "growth" in label or "margin" in label:
            row[1].number_format = "0.0%"
        elif isinstance(row[1].value, float):
            row[1].number_format = "0.00"
    style_sheet(summary)

    for name in ("bear", "base", "bull"):
        sheet = wb.create_sheet(name.title())
        sheet.append(["Year", "Revenue growth", "Revenue ($B)", "FCF margin", "Restructuring cash ($B)", "FCF ($B)", "Diluted shares (B)", "Discount period", "PV/share"])
        for row in payload["scenarios"][name]["rows"]:
            sheet.append([
                row["year"], row["revenue_growth"], row["revenue_b"], row["fcf_margin"],
                row["restructuring_cash_b"], row["fcf_b"], row["diluted_shares_b"],
                row["discount_period"], row["pv_per_share"],
            ])
        for row in sheet.iter_rows(min_row=2):
            row[1].number_format = "0.0%"
            row[3].number_format = "0.0%"
            for index in (2, 4, 5, 6, 7, 8):
                row[index].number_format = "0.000"
        style_sheet(sheet)

    sensitivity = wb.create_sheet("Sensitivity")
    terminal_rates = [0.02, 0.025, 0.03, 0.035, 0.04]
    sensitivity.append(["Discount / terminal growth"] + terminal_rates)
    for discount_rate in [0.095, 0.10, 0.105, 0.11, 0.115, 0.12, 0.125]:
        sensitivity.append([discount_rate] + [payload["sensitivity"][f"{discount_rate:.3f}"][f"{growth:.3f}"] for growth in terminal_rates])
    for row in sensitivity.iter_rows(min_row=2):
        row[0].number_format = "0.0%"
        for cell in row[1:]:
            cell.number_format = "$0.00"
    for cell in sensitivity[1][1:]:
        cell.number_format = "0.0%"
    style_sheet(sensitivity)

    reverse = wb.create_sheet("Reverse DCF")
    reverse.append(["Test", "Result", "Held fixed"])
    reverse.append(["Required constant revenue growth, 2027-35", payload["reverse_dcf_growth"]["constant_revenue_growth_2027_2035"], payload["reverse_dcf_growth"]["held_fixed"]])
    reverse.append(["Implied 2035 revenue ($B)", payload["reverse_dcf_growth"]["implied_2035_revenue_b"], "Same growth solve"])
    reverse.append(["Implied 2035 FCF ($B)", payload["reverse_dcf_growth"]["implied_2035_fcf_b"], "Same growth solve"])
    reverse.append(["Required terminal FCF margin", payload["reverse_dcf_margin"]["terminal_fcf_margin"], payload["reverse_dcf_margin"]["held_fixed"]])
    reverse.append(["Heroic case fair value", payload["heroic_market_case"]["fair_value_per_share"], "30% annual growth; 40% terminal FCF margin; 10% discount; 3.5% terminal growth"])
    reverse["B2"].number_format = "0.0%"
    reverse["B5"].number_format = "0.0%"
    reverse["B6"].number_format = "$0.00"
    style_sheet(reverse)

    sources = wb.create_sheet("Sources")
    sources.append(["Source", "URL", "Used for"])
    for source in SOURCES:
        sources.append([source["label"], source["url"], source["use"]])
    for limitation in payload["limitations"]:
        sources.append(["Limitation", "", limitation])
    style_sheet(sources)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUT)


def verify(payload: dict) -> None:
    expected = {"bear": 37.49, "base": 60.70, "bull": 108.06}
    for name, target in expected.items():
        actual = payload["scenarios"][name]["fair_value_per_share"]
        if not math.isclose(actual, target, abs_tol=0.02):
            raise ValueError(f"{name} fair value drifted: {actual:.2f} != {target:.2f}")
    if payload["reverse_dcf_growth"]["constant_revenue_growth_2027_2035"] < 0.40:
        raise ValueError("Reverse DCF growth hurdle unexpectedly fell below 40%")
    if not zipfile.is_zipfile(XLSX_OUT):
        raise ValueError("DCF workbook is not a valid XLSX archive")
    with zipfile.ZipFile(XLSX_OUT) as workbook:
        names = set(workbook.namelist())
        expected_sheets = {f"xl/worksheets/sheet{index}.xml" for index in range(1, 8)}
        if not expected_sheets <= names:
            raise ValueError("DCF workbook is missing one or more worksheets")
        workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8")
        for sheet in ("Summary", "Bear", "Base", "Bull", "Sensitivity", "Reverse DCF", "Sources"):
            if f'name="{sheet}"' not in workbook_xml:
                raise ValueError(f"DCF workbook is missing sheet {sheet}")
    if len(SOURCES) < 4:
        raise ValueError("DCF needs at least four source receipts")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    payload = build_payload()
    if args.check:
        checked = json.loads(JSON_OUT.read_text())
        if checked != payload:
            raise SystemExit("NET DCF JSON is stale; run scripts/build-net-dcf.py")
        verify(checked)
        print("[net-dcf] current: bear $37.49 / base $60.70 / bull $108.06")
        return 0
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, indent=2) + "\n")
    build_workbook(payload)
    verify(payload)
    print(f"[net-dcf] built {JSON_OUT.relative_to(ROOT)} and {XLSX_OUT.relative_to(ROOT)}")
    print("[net-dcf] fair values: bear $37.49 / base $60.70 / bull $108.06")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
