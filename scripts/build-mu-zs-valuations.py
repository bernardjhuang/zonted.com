#!/usr/bin/env python3
"""Build filing-backed valuation receipts for MU and ZS theses."""
from __future__ import annotations

import argparse
import json
import math
import zipfile
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "trading" / "research"
DATE = "2026-08-07"

MU = {
    "symbol": "MU",
    "valuation_date": DATE,
    "quote": 877.56,
    "quote_timestamp": "2026-08-07T23:00:00Z",
    "method": "cycle-normalized GAAP earnings-power scenarios",
    "inputs": {
        "fq3_2026_revenue_b": 41.456,
        "fq3_2026_gaap_eps": 24.67,
        "fq3_2026_adjusted_fcf_b": 18.3,
        "fq4_2026_revenue_guide_b": 50.0,
        "fq4_2026_gaap_eps_guide": 30.73,
        "cash_investments_restricted_cash_b": 30.2,
        "long_term_debt_b": 5.140,
    },
    "scenarios": {
        "bear": {"normalized_eps": 30.0, "earnings_multiple": 10.0},
        "base": {"normalized_eps": 50.0, "earnings_multiple": 14.0},
        "bull": {"normalized_eps": 80.0, "earnings_multiple": 16.0},
    },
    "sources": [
        {
            "label": "Micron fiscal Q3 2026 earnings release",
            "url": "https://www.sec.gov/Archives/edgar/data/723125/000072312526000013/a2026q3ex991-pressrelease.htm",
            "use": "Q3 results, adjusted FCF, Q4 revenue/EPS outlook, HBM4 milestones",
        },
        {
            "label": "Micron fiscal Q3 2026 Form 10-Q",
            "url": "https://www.sec.gov/Archives/edgar/data/723125/000072312526000015/mu-20260528.htm",
            "use": "Cash, securities, debt, capex and strategic customer agreement terms",
        },
        {
            "label": "Robinhood MU quote and fundamentals",
            "url": "https://robinhood.com/stocks/MU",
            "use": "August 7 quote and market data",
        },
    ],
    "limitations": [
        "Micron is at an extreme memory-cycle peak; applying peak quarterly economics to a perpetuity would be fake precision.",
        "The normalized EPS cases are judgmental and must be reset when DRAM/NAND pricing, HBM mix or FY2027 guidance changes.",
        "Scenario values are valuation references, not automatic buy or sell targets.",
    ],
}
for spec in MU["scenarios"].values():
    spec["fair_value_per_share"] = spec["normalized_eps"] * spec["earnings_multiple"]
    spec["upside_downside_to_quote"] = spec["fair_value_per_share"] / MU["quote"] - 1
MU["reverse_earnings_power"] = {
    "normalized_eps_required_at_base_multiple": MU["quote"] / MU["scenarios"]["base"]["earnings_multiple"],
    "held_fixed": "14x cycle-normalized GAAP earnings multiple",
}

ZS_SCENARIOS = {
    "bear": {
        "revenue_growth": [0.16, 0.14, 0.12, 0.10, 0.08, 0.06, 0.05, 0.04, 0.03],
        "fcf_margin": [0.21, 0.22, 0.23, 0.24, 0.25, 0.26, 0.26, 0.26, 0.26],
        "discount_rate": 0.125,
        "terminal_growth": 0.025,
        "annual_dilution": 0.03,
    },
    "base": {
        "revenue_growth": [0.22, 0.20, 0.18, 0.16, 0.14, 0.12, 0.10, 0.08, 0.06],
        "fcf_margin": [0.23, 0.24, 0.25, 0.26, 0.28, 0.29, 0.30, 0.31, 0.31],
        "discount_rate": 0.11,
        "terminal_growth": 0.03,
        "annual_dilution": 0.02,
    },
    "bull": {
        "revenue_growth": [0.25, 0.23, 0.21, 0.19, 0.17, 0.15, 0.13, 0.10, 0.08],
        "fcf_margin": [0.24, 0.25, 0.27, 0.29, 0.31, 0.33, 0.34, 0.35, 0.35],
        "discount_rate": 0.095,
        "terminal_growth": 0.035,
        "annual_dilution": 0.015,
    },
}

ZS = {
    "symbol": "ZS",
    "valuation_date": DATE,
    "quote": 168.70,
    "quote_timestamp": "2026-08-07T23:00:00Z",
    "method": "SEC-first 9-year diluted per-share FCF DCF",
    "inputs": {
        "fy2026_revenue_guide_midpoint_b": 3.331,
        "fy2026_fcf_margin_guide_midpoint": 0.2305,
        "cash_and_short_term_investments_b": 3.539107,
        "convertible_debt_principal_b": 1.725,
        "net_cash_b": 1.814107,
        "diluted_shares_b": 0.168,
        "nine_month_stock_compensation_b": 0.610332,
        "remaining_performance_obligations_b": 6.4593,
    },
    "scenarios": {},
    "sources": [
        {
            "label": "Zscaler fiscal Q3 2026 earnings release",
            "url": "https://www.sec.gov/Archives/edgar/data/1713683/000171368326000095/zs-04302026_991.htm",
            "use": "ARR, revenue, billings, FCF and FY2026 guidance",
        },
        {
            "label": "Zscaler fiscal Q3 2026 Form 10-Q",
            "url": "https://www.sec.gov/Archives/edgar/data/1713683/000171368326000096/zs-20260430.htm",
            "use": "Cash, investments, converts, RPO, shares and stock compensation",
        },
        {
            "label": "Robinhood ZS quote and fundamentals",
            "url": "https://robinhood.com/stocks/ZS",
            "use": "August 7 quote and market data",
        },
    ],
    "limitations": [
        "FCF includes a large stock-compensation add-back, so every case explicitly grows the diluted share count.",
        "Terminal value is material; the bear/base/bull spread is the honest output, not noise to average away.",
        "The FY2026 FCF-margin guide fell after higher capex; margin expansion must be earned rather than assumed.",
    ],
}


def zs_forecast(name: str, spec: dict | None = None) -> dict:
    spec = dict(spec or ZS_SCENARIOS[name])
    revenue = ZS["inputs"]["fy2026_revenue_guide_midpoint_b"]
    shares = ZS["inputs"]["diluted_shares_b"]
    rows = []
    explicit = 0.0
    for index, year in enumerate(range(2027, 2036), 1):
        growth = spec["revenue_growth"][index - 1]
        margin = spec["fcf_margin"][index - 1]
        revenue *= 1 + growth
        fcf = revenue * margin
        diluted = shares * (1 + spec["annual_dilution"]) ** index
        pv = (fcf / diluted) / (1 + spec["discount_rate"]) ** index
        explicit += pv
        rows.append({
            "year": year,
            "revenue_growth": growth,
            "revenue_b": revenue,
            "fcf_margin": margin,
            "fcf_b": fcf,
            "diluted_shares_b": diluted,
            "pv_per_share": pv,
        })
    last = rows[-1]
    terminal = (
        last["fcf_b"] * (1 + spec["terminal_growth"])
        / (spec["discount_rate"] - spec["terminal_growth"])
        / last["diluted_shares_b"]
        / (1 + spec["discount_rate"]) ** len(rows)
    )
    net_cash = ZS["inputs"]["net_cash_b"] / shares
    fair = explicit + terminal + net_cash
    return {
        "name": name,
        "assumptions": spec,
        "rows": rows,
        "explicit_pv_per_share": explicit,
        "terminal_pv_per_share": terminal,
        "net_cash_per_share": net_cash,
        "fair_value_per_share": fair,
        "terminal_value_share_of_operating_value": terminal / (explicit + terminal),
        "upside_downside_to_quote": fair / ZS["quote"] - 1,
    }


for case in ("bear", "base", "bull"):
    ZS["scenarios"][case] = zs_forecast(case)


def solve_reverse_growth() -> dict:
    low, high = 0.0, 0.60
    base = ZS_SCENARIOS["base"]
    for _ in range(100):
        growth = (low + high) / 2
        spec = dict(base)
        spec["revenue_growth"] = [growth] * 9
        value = zs_forecast("reverse growth", spec)["fair_value_per_share"]
        if value < ZS["quote"]:
            low = growth
        else:
            high = growth
    growth = (low + high) / 2
    spec = dict(base)
    spec["revenue_growth"] = [growth] * 9
    result = zs_forecast("reverse growth", spec)
    return {
        "constant_revenue_growth_2027_2035": growth,
        "implied_2035_revenue_b": result["rows"][-1]["revenue_b"],
        "matched_value_per_share": result["fair_value_per_share"],
        "held_fixed": "base FCF-margin path ending at 31%; 11% discount rate; 3% terminal growth; 2% annual dilution",
    }


def solve_reverse_margin() -> dict:
    low, high = 0.10, 0.80
    base = ZS_SCENARIOS["base"]
    for _ in range(100):
        margin = (low + high) / 2
        spec = dict(base)
        spec["fcf_margin"] = [0.23 + (margin - 0.23) * i / 8 for i in range(9)]
        value = zs_forecast("reverse margin", spec)["fair_value_per_share"]
        if value < ZS["quote"]:
            low = margin
        else:
            high = margin
    margin = (low + high) / 2
    spec = dict(base)
    spec["fcf_margin"] = [0.23 + (margin - 0.23) * i / 8 for i in range(9)]
    result = zs_forecast("reverse margin", spec)
    return {
        "terminal_fcf_margin": margin,
        "matched_value_per_share": result["fair_value_per_share"],
        "held_fixed": "base revenue-growth path; 11% discount rate; 3% terminal growth; 2% annual dilution",
    }


ZS["reverse_dcf_growth"] = solve_reverse_growth()
ZS["reverse_dcf_margin"] = solve_reverse_margin()


def clean(value: Any) -> Any:
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("Non-finite value")
        return round(value, 12)
    if isinstance(value, dict):
        return {key: clean(item) for key, item in value.items()}
    if isinstance(value, list):
        return [clean(item) for item in value]
    return value


def style_sheet(ws):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 52)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def build_workbook(payload: dict, path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)
    summary = wb.create_sheet("Summary")
    summary.append(["Field", "Value"])
    summary.append(["Symbol", payload["symbol"]])
    summary.append(["Valuation date", payload["valuation_date"]])
    summary.append(["Quote", payload["quote"]])
    summary.append(["Method", payload["method"]])
    for case in ("bear", "base", "bull"):
        summary.append([f"{case.title()} value", payload["scenarios"][case]["fair_value_per_share"]])
    summary.append(["Important", "Scenario references only; read the receipt limitations."])
    style_sheet(summary)

    inputs = wb.create_sheet("Inputs")
    inputs.append(["Input", "Value"])
    for key, value in payload["inputs"].items():
        inputs.append([key, value])
    style_sheet(inputs)

    for case in ("bear", "base", "bull"):
        ws = wb.create_sheet(case.title())
        scenario = payload["scenarios"][case]
        if payload["symbol"] == "MU":
            ws.append(["Normalized EPS", "Multiple", "Fair value", "Upside / downside"])
            ws.append([
                scenario["normalized_eps"],
                scenario["earnings_multiple"],
                "=A2*B2",
                f"=C2/{payload['quote']}-1",
            ])
        else:
            ws.append(["Year", "Revenue growth", "Revenue ($B)", "FCF margin", "FCF ($B)", "Diluted shares (B)", "PV/share"])
            for row in scenario["rows"]:
                ws.append([row["year"], row["revenue_growth"], row["revenue_b"], row["fcf_margin"], row["fcf_b"], row["diluted_shares_b"], row["pv_per_share"]])
            ws.append([])
            ws.append(["Explicit PV/share", scenario["explicit_pv_per_share"]])
            ws.append(["Terminal PV/share", scenario["terminal_pv_per_share"]])
            ws.append(["Net cash/share", scenario["net_cash_per_share"]])
            ws.append(["Fair value/share", scenario["fair_value_per_share"]])
        style_sheet(ws)

    sources = wb.create_sheet("Sources")
    sources.append(["Label", "URL", "Use"])
    for source in payload["sources"]:
        sources.append([source["label"], source["url"], source["use"]])
        sources.cell(sources.max_row, 2).hyperlink = source["url"]
        sources.cell(sources.max_row, 2).style = "Hyperlink"
    style_sheet(sources)

    limitations = wb.create_sheet("Limitations")
    limitations.append(["Model limitation"])
    for limitation in payload["limitations"]:
        limitations.append([limitation])
    style_sheet(limitations)
    wb.save(path)


def outputs(payload: dict) -> tuple[Path, Path]:
    slug = "mu-normalized-earnings" if payload["symbol"] == "MU" else "zs-dcf"
    return OUT / f"{slug}-{DATE}.json", OUT / f"{slug}-{DATE}.xlsx"


def write_all() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    for payload in (clean(MU), clean(ZS)):
        json_path, xlsx_path = outputs(payload)
        json_path.write_text(json.dumps(payload, indent=2) + "\n")
        build_workbook(payload, xlsx_path)
        print(f"wrote {json_path.relative_to(ROOT)} and {xlsx_path.relative_to(ROOT)}")


def check_all() -> None:
    for payload in (clean(MU), clean(ZS)):
        json_path, xlsx_path = outputs(payload)
        if json.loads(json_path.read_text()) != payload:
            raise SystemExit(f"stale receipt: {json_path}")
        if xlsx_path.stat().st_size < 10_000 or not zipfile.is_zipfile(xlsx_path):
            raise SystemExit(f"invalid workbook: {xlsx_path}")
        with zipfile.ZipFile(xlsx_path) as workbook:
            workbook_xml = workbook.read("xl/workbook.xml").decode("utf-8")
            expected = {"Summary", "Inputs", "Bear", "Base", "Bull", "Sources", "Limitations"}
            for sheet in expected:
                if f'name="{sheet}"' not in workbook_xml:
                    raise SystemExit(f"workbook missing {sheet}: {xlsx_path}")
            if payload["symbol"] == "MU":
                base_xml = workbook.read("xl/worksheets/sheet4.xml").decode("utf-8")
                if "<f>A2*B2</f>" not in base_xml:
                    raise SystemExit("MU workbook lost its valuation formula")
    print("MU and ZS valuation receipts are current")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    check_all() if args.check else write_all()


if __name__ == "__main__":
    main()
