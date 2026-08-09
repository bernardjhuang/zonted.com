#!/usr/bin/env python3
"""Build filing-backed valuation receipts for Salesforce and CoreWeave."""
from __future__ import annotations

import argparse
import json
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "trading" / "research"
DATE = "2026-08-07"
SEC_CRM_RELEASE = "https://www.sec.gov/Archives/edgar/data/1108524/000110852426000125/crm-q1fy27xexhibit991.htm"
SEC_CRM_10Q = "https://www.sec.gov/Archives/edgar/data/1108524/000110852426000127/crm-20260430.htm"
SEC_CRWV_RELEASE = "https://www.sec.gov/Archives/edgar/data/1769628/000176962826000220/coreweave1q26earningspress.htm"
SEC_CRWV_10Q = "https://www.sec.gov/Archives/edgar/data/1769628/000176962826000222/crwv-20260331.htm"


def crm_payload() -> dict:
    price = 192.73
    scenarios = {
        "bear": {"normalized_eps": 7.96, "multiple": 18.0},
        "base": {"normalized_eps": 11.00, "multiple": 20.0},
        "bull": {"normalized_eps": 14.09, "multiple": 22.0},
    }
    for row in scenarios.values():
        row["fair_value_per_share"] = round(row["normalized_eps"] * row["multiple"], 2)
    return {
        "symbol": "CRM",
        "company": "Salesforce, Inc.",
        "valuation_date": DATE,
        "quote_usd": price,
        "method": "blended normalized EPS scenarios",
        "market_snapshot": {"market_cap_usd_b": 158.54, "shares_outstanding_m": 822.61, "trailing_pe": 21.52},
        "filing_inputs": {
            "q1_fy27_revenue_usd_b": 11.133,
            "q1_fy27_crpo_usd_b": 33.6,
            "q1_fy27_gaap_operating_margin_pct": 21.1,
            "q1_fy27_non_gaap_operating_margin_pct": 34.8,
            "fy27_revenue_guide_midpoint_usd_b": 46.05,
            "fy27_gaap_eps_guide_midpoint": 7.96,
            "fy27_non_gaap_eps_guide_midpoint": 14.09,
            "fy27_non_gaap_operating_margin_guide_pct": 34.3,
        },
        "scenarios": scenarios,
        "market_implied_normalized_eps_at_base_multiple": round(price / 20.0, 2),
        "sources": [SEC_CRM_RELEASE, SEC_CRM_10Q],
        "limitations": [
            "The scenario range intentionally spans GAAP-like to company non-GAAP earnings definitions; stock compensation and acquisition adjustments remain economically material.",
            "The $25 billion accelerated repurchase was debt-funded, so lower share count must be judged against higher interest expense and leverage.",
            "This static receipt is not a forecast, price target, or investment recommendation.",
        ],
    }


def crwv_payload() -> dict:
    price = 90.67
    market_cap = 49.46686336249
    shares_b = 0.545570347
    net_debt = 24.859 - 2.244 - 0.022
    scenarios = {
        "bear": {"normalized_revenue_usd_b": 10.0, "ev_revenue_multiple": 4.0},
        "base": {"normalized_revenue_usd_b": 12.0, "ev_revenue_multiple": 6.0},
        "bull": {"normalized_revenue_usd_b": 14.0, "ev_revenue_multiple": 8.0},
    }
    for row in scenarios.values():
        enterprise_value = row["normalized_revenue_usd_b"] * row["ev_revenue_multiple"]
        row["enterprise_value_usd_b"] = round(enterprise_value, 3)
        row["equity_value_usd_b"] = round(enterprise_value - net_debt, 3)
        row["fair_value_per_share"] = round((enterprise_value - net_debt) / shares_b, 2)
    enterprise_value = market_cap + net_debt
    return {
        "symbol": "CRWV",
        "company": "CoreWeave, Inc.",
        "valuation_date": DATE,
        "quote_usd": price,
        "method": "net-debt-adjusted normalized revenue scenarios",
        "market_snapshot": {
            "market_cap_usd_b": round(market_cap, 3),
            "shares_outstanding_m": round(shares_b * 1000, 2),
            "debt_usd_b": 24.859,
            "cash_usd_b": 2.244,
            "marketable_securities_usd_b": 0.022,
            "net_debt_usd_b": round(net_debt, 3),
            "enterprise_value_usd_b": round(enterprise_value, 3),
        },
        "filing_inputs": {
            "q1_2026_revenue_usd_b": 2.078,
            "q1_2026_adjusted_ebitda_usd_b": 1.157,
            "q1_2026_adjusted_ebitda_margin_pct": 56.0,
            "q1_2026_operating_loss_usd_b": 0.144,
            "q1_2026_interest_expense_usd_b": 0.536,
            "q1_2026_revenue_backlog_usd_b": 99.4,
            "top_two_customer_revenue_pct": 65.0,
        },
        "scenarios": scenarios,
        "market_implied_normalized_revenue_at_base_multiple_usd_b": round(enterprise_value / 6.0, 2),
        "sources": [SEC_CRWV_RELEASE, SEC_CRWV_10Q],
        "limitations": [
            "Revenue backlog includes committed contracts subject to delivery, service-availability, power, construction, and customer-performance conditions; it is not cash in hand.",
            "Enterprise value subtracts reported cash and marketable securities from reported debt but omits restricted cash and other financing-like obligations.",
            "Revenue multiples are highly sensitive to utilization, GPU economics, financing costs, capex execution, and customer concentration.",
            "This static receipt is not a forecast, price target, or investment recommendation.",
        ],
    }


def style_sheet(ws, header_fill: str, widths: tuple[int, int] = (34, 76)) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = widths[0]
    ws.column_dimensions["B"].width = widths[1]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_pairs(ws, pairs: list[tuple[str, object]]) -> None:
    for key, value in pairs:
        ws.append([key, value])


def build_workbook(payload: dict, path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    values = payload["scenarios"]
    add_pairs(summary, [
        ("Symbol", payload["symbol"]), ("Company", payload["company"]),
        ("Valuation date", payload["valuation_date"]), ("Quote (USD)", payload["quote_usd"]),
        ("Method", payload["method"]),
        ("Bear fair value", values["bear"]["fair_value_per_share"]),
        ("Base fair value", values["base"]["fair_value_per_share"]),
        ("Bull fair value", values["bull"]["fair_value_per_share"]),
    ])
    style_sheet(summary, "17324D")

    inputs = wb.create_sheet("Inputs")
    inputs.append(["Input", "Value"])
    add_pairs(inputs, [(k, v) for group in (payload["market_snapshot"], payload["filing_inputs"]) for k, v in group.items()])
    style_sheet(inputs, "0B6E4F")

    methodology = wb.create_sheet("Methodology")
    methodology.append(["Step", "Description"])
    methodology.append([1, f"Start from the {payload['valuation_date']} market snapshot and official filing inputs."])
    methodology.append([2, f"Apply the stated {payload['method']} without treating the output as a forecast."])
    methodology.append([3, "Use explicit bear, base, and bull assumptions rather than a single-point target."])
    methodology.append([4, "Compare the range with the quoted market price and disclosed balance-sheet constraints."])
    methodology.append([5, "Rebuild after the next earnings release because inputs and capital structure can change materially."])
    style_sheet(methodology, "334155", (12, 110))

    for name, row in values.items():
        ws = wb.create_sheet(name.title())
        ws.append(["Scenario field", "Value"])
        add_pairs(ws, list(row.items()))
        style_sheet(ws, {"bear": "8A2D2D", "base": "22577A", "bull": "256D3A"}[name])

    sources = wb.create_sheet("Sources")
    sources.append(["Type", "URL"])
    for url in payload["sources"]:
        sources.append(["Official SEC filing/exhibit", url])
    style_sheet(sources, "4B5563", (28, 110))

    limits = wb.create_sheet("Limitations")
    limits.append(["Item", "Text"])
    for i, text in enumerate(payload["limitations"], 1):
        limits.append([i, text])
    style_sheet(limits, "6B4F1D", (10, 120))
    wb.save(path)


def validate(path: Path, symbol: str) -> None:
    if not path.exists() or path.stat().st_size < 10_000:
        raise SystemExit(f"{path} is missing or too small")
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        if "xl/workbook.xml" not in names or len([n for n in names if n.startswith("xl/worksheets/sheet")]) < 8:
            raise SystemExit(f"{path} is not a complete workbook")
        shared = archive.read("xl/sharedStrings.xml").decode("utf-8", "ignore") if "xl/sharedStrings.xml" in names else ""
        if shared and symbol not in shared:
            raise SystemExit(f"{path} does not contain {symbol}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()
    models = {"crm": crm_payload(), "crwv": crwv_payload()}
    for slug, payload in models.items():
        json_path = OUT / f"{slug}-valuation-{DATE}.json"
        xlsx_path = OUT / f"{slug}-valuation-{DATE}.xlsx"
        if not args.check:
            json_path.write_text(json.dumps(payload, indent=2) + "\n")
            build_workbook(payload, xlsx_path)
        if not json_path.exists():
            raise SystemExit(f"{json_path} is missing")
        saved = json.loads(json_path.read_text())
        if saved["valuation_date"] != DATE or saved["symbol"] != payload["symbol"]:
            raise SystemExit(f"{json_path} is stale")
        validate(xlsx_path, payload["symbol"])
        print(f"validated {xlsx_path.relative_to(ROOT)} ({xlsx_path.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
